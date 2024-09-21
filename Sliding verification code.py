from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
import time
import re
import os
from urllib.request import urlretrieve
import cv2
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook


def find_gap_position(image_path):
    # 读取图片
    img = cv2.imread(image_path)
    # 转换为灰度图
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # 使用高斯模糊减少噪声
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    # 使用 Canny 边缘检测
    edges = cv2.Canny(blurred, 50, 150)

    # 寻找轮廓
    contours, _ = cv2.findContours(edges, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    # 寻找最大的轮廓
    if len(contours) > 0:
        largest_contour = max(contours, key=cv2.contourArea)
        x, y, w, h = cv2.boundingRect(largest_contour)

        # 返回缺口最左侧的位置
        return x  # 缺口最左侧的位置

    else:
        return None


# Firefox WebDriver 的路径
geckodriver_path = r"D:\geckodriver\geckodriver.exe"  # 替换为Firefox WebDriver 的路径

# 设置 Firefox WebDriver
service = Service(executable_path=geckodriver_path)
options = webdriver.FirefoxOptions()
options.add_argument("--start-maximized")  # 启动时最大化浏览器窗口

driver = webdriver.Firefox(service=service, options=options)

try:
    url = r""  # 所要访问网站的地址
    driver.get(url)
    time.sleep(2)

    # 点击验证码按钮
    captcha_button = driver.find_element(By.XPATH, r'//*[@id="TencentCaptcha"]')
    captcha_button.click()
    time.sleep(2)

    # 切换到验证码 iframe
    captcha_frame = driver.find_element(By.ID, "tcaptcha_iframe_dy")
    driver.switch_to.frame(captcha_frame)

    # 获取背景图片 URL
    background_image = driver.find_element(By.ID, "slideBg").value_of_css_property(
        "background-image"
    )
    var = re.findall(r'url\("(.+)"\)', background_image)[0]
    urlretrieve(var, "beijing.png")
    # 使用函数找到缺口位置
    gap_position = find_gap_position("beijing.png")

    if gap_position is not None:
        print(f"下载图片缺口位置: {gap_position}")
    else:
        print("未找到缺口")
        raise Exception("未能找到缺口位置")

    real_gap_position = 811 / 672 * gap_position
    print("真正的距离：", real_gap_position)
    move_space = int((340 / 811 * real_gap_position) - 34)

    slider = driver.find_element(By.XPATH, r"/html/body/div/div[3]/div[2]/div[6]")
    # 初始化动作链
    action_chains = ActionChains(driver)

    # 模拟人类的移动
    action_chains.click_and_hold(slider).perform()

    # 直接移动到目标位置
    action_chains.move_by_offset(int(move_space), 0).perform()

    # 释放鼠标
    action_chains.release().perform()

finally:
    # 删除临时截图文件
    try:
        os.remove("beijing.png")
    except FileNotFoundError:
        print("文件可能已被删除或不存在")

# 切换回主文档
driver.switch_to.default_content()

# 检查 Excel 文件是否存在
excel_file = ".xlsx"  # 替换为要保存的文件名
if os.path.exists(excel_file):
    workbook = load_workbook(excel_file)
    worksheet = workbook.active
else:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "test_1"
""" 表头可以自定义 """
# 检查是否已有表头
if worksheet["A1"].value != "院校名称":
    # 如果没有表头，写入表头
    worksheet["A1"] = "院校名称"
    worksheet["B1"] = "总分"
    worksheet["C1"] = "语文"
    worksheet["D1"] = "数学"
    worksheet["E1"] = "外语听力"

# 填写查询条件并提交
query_input = WebDriverWait(driver, 10).until(
    ec.presence_of_element_located((By.XPATH, '//*[@id="querytxt"]'))
)
query_input.send_keys("0001")
time.sleep(1)  # 等待片刻以确保输入生效
query_button = driver.find_element(By.XPATH, r'//*[@id="QueryBtn"]')
query_button.click()

# 等待查询结果页面加载完成
WebDriverWait(driver, 10).until(ec.presence_of_element_located((By.ID, "tabInfo")))

html_content = driver.page_source  # 获取当前页面的 HTML
# 解析 HTML
soup = BeautifulSoup(html_content, "lxml")

# 查找包含数据的表格
table = soup.find("table", id="tabInfo")

# 提取所需的数据
data_rows = table.find_all("tr")[1:]  # 忽略表头行

# 找到最近的一个空行
last_row_with_data = worksheet.max_row
while last_row_with_data > 1 and all(
    worksheet.cell(row=last_row_with_data, column=col).value is None
    for col in range(1, 6)
):
    last_row_with_data -= 1
row_index = last_row_with_data + 1

for row in data_rows:
    cells = row.find_all("td")
    if len(cells) >= 6:  # 确保列表至少有6个元素
        institution_name = cells[1].get_text(strip=True)
        total_score = cells[2].get_text(strip=True)
        chinese_score = cells[3].get_text(strip=True)
        math_score = cells[4].get_text(strip=True)
        english_score = cells[5].get_text(strip=True)

        # 在 Excel 中写入数据
        worksheet[f"A{row_index}"] = institution_name
        worksheet[f"B{row_index}"] = total_score
        worksheet[f"C{row_index}"] = chinese_score
        worksheet[f"D{row_index}"] = math_score
        worksheet[f"E{row_index}"] = english_score

        # 输出数据
        print(
            f"{institution_name} {total_score} {chinese_score} {math_score} {english_score}"
        )

        row_index += 1
    else:
        print("跳过此行，因为它缺少必要的数据。")

# 保存 Excel 文件
workbook.save(excel_file)

# 关闭浏览器
driver.quit()
